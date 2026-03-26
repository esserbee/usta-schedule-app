from flask import Flask, request, send_file, render_template_string
import io
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin
import json

app = Flask(__name__)

HTML_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>USTA NorCal League Schedule Organizer</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body { font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; padding: 2rem; max-width: 1000px; margin: 0 auto; background: #f7f6f2; color: #222; }
    h1 { font-size: 1.9rem; margin-bottom: 0.5rem; }
    p { max-width: 70ch; }
    label { font-weight: 600; display: block; margin-bottom: 0.25rem; }
    textarea { width: 100%; min-height: 140px; padding: 0.75rem; border-radius: 8px; border: 1px solid #ccc; font-family: monospace; font-size: 0.9rem; }
    input[type="url"] { width: 100%; padding: 0.6rem 0.75rem; border-radius: 8px; border: 1px solid #ccc; font-size: 0.9rem; margin-bottom: 0.5rem; }
    button { margin-top: 1rem; padding: 0.6rem 1.2rem; border-radius: 999px; border: none; background: #01696f; color: #fff; font-weight: 600; cursor: pointer; }
    button:hover { background: #0c4e54; }
    .help { font-size: 0.85rem; color: #666; margin-top: 0.25rem; }
    .status { margin-top: 1rem; font-size: 0.9rem; color: #444; }
    .error { color: #a12c2c; }
    code { background: #f0efea; padding: 0.1rem 0.3rem; border-radius: 4px; }
    .results { margin-top: 2rem; }
    table { width: 100%; border-collapse: collapse; margin-top: 1rem; font-size: 0.9rem; }
    th, td { border: 1px solid #ddd; padding: 0.4rem 0.6rem; text-align: left; }
    th { background: #01696f; color: #fff; font-weight: 700; }
    tr.conflict { background: #fff7c2; }
    tr.pending { background: #efefef; color: #9a9a9a; font-style: italic; }
    tr.conflict.pending { background: #efefef !important; color: #9a9a9a !important; font-style: italic; }
    tr.pending td { color: #9a9a9a !important; font-style: italic !important; }
    tr.conflict.pending td { color: #9a9a9a !important; font-style: italic !important; }
    tr.conflict td { color: #222; }
    td.homeaway_home { color: #01696f !important; font-weight: 700; }
    td.homeaway_away { color: #a12c2c !important; font-weight: 700; }
    td.location_home { color: #0b6b0b !important; font-weight: 600; }
    td.location_away { color: #6b2d91 !important; font-weight: 600; }
    .footnote-link { margin-left: 0.25rem; color: #01696f; font-weight: 900; text-decoration: none; }
    .footnote-link:hover { text-decoration: underline; }
    .actions { margin-top: 1rem; }
    .team-list label { font-weight: 400; }
    fieldset { border: 1px solid #ddd; padding: 0.75rem 1rem; border-radius: 8px; margin-bottom: 1rem; }
    legend { padding: 0 0.25rem; font-weight: 600; }
  </style>
</head>
<body>
  <h1>USTA NorCal League Schedule Organizer</h1>
  <p>Step 1: Choose whether to use your USTA NorCal player profile or paste individual team info URLs.</p>
  <p>Step 2: Review the combined schedule and download an Excel file with conflicts highlighted.</p>

  <form method="post" action="/generate">
    <fieldset>
      <legend>Step 1: Input method</legend>
      <label>
        <input type="radio" name="mode" value="profile" {% if mode == 'profile' %}checked{% endif %}>
        Use player profile URL
      </label>
      <label>
        <input type="radio" name="mode" value="teams" {% if mode == 'teams' %}checked{% endif %}>
        Use individual team info URLs
      </label>
    </fieldset>

    <div id="profile_url_wrap" {% if mode != 'profile' %}style="display:none"{% endif %}>
      <label for="profile_url">Player profile URL</label>
      <input type="url" id="profile_url" name="profile_url" placeholder="https://leagues.ustanorcal.com/...playermatches.asp?id=..." value="{{ profile_url or '' }}">
      <div class="help">Used when "Use player profile URL" is selected.</div>
    </div>

    <div id="urls_wrap" {% if mode != 'teams' %}style="display:none"{% endif %}>
      <label for="urls">Team info URLs</label>
      <textarea id="urls" name="urls" placeholder="https://leagues.ustanorcal.com/teaminfo.asp?id=109510
https://leagues.ustanorcal.com/teaminfo.asp?id=109621">{{ urls_value or '' }}</textarea>
      <div class="help">Used when "Use individual team info URLs" is selected (one URL per line).</div>
    </div>

    <button type="submit">Next</button>
  </form>

  {% if message %}
  <div class="status {% if error %}error{% endif %}">{{ message }}</div>
  {% endif %}

  {% if team_choices %}
  <div class="results">
    <h2>Select teams for {{ current_year }}</h2>
    {% if player_first_name %}
      <p class="help">Hi {{ player_first_name }} — select the teams below to build your schedule.</p>
    {% endif %}
    {% if filtered_to_year %}
      <p>Showing teams detected for {{ current_year }} from your profile. Select which ones to include in the schedule.</p>
    {% else %}
      <p>These are the teams found on your profile. Select which ones to include in the schedule.</p>
    {% endif %}
    <form method="post" action="/generate">
      <input type="hidden" name="profile_url" value="{{ profile_url | e }}">
      <input type="hidden" name="mode" value="profile">
      <input type="hidden" name="player_name" value="{{ player_name | e }}">
      <div class="team-list">
      {% for t in team_choices %}
        <div>
          <label>
            <input type="checkbox" name="team_urls" value="{{ t.url }}" checked>
            {{ t.label }}
          </label>
        </div>
      {% endfor %}
      </div>
      <button type="submit">Build Schedule from Selected Teams</button>
    </form>
  </div>
  {% endif %}

  {% if schedule %}
  <div class="results">
    <h2>Combined schedule</h2>
    <p>Rows highlighted in yellow are days where you have more than one match scheduled.</p>
    <p id="location-footnote" style="font-size: 1.05rem; color: #a12c2c; font-weight: 600;">
      * Entry in the location column may not be accurate sometimes, use &ldquo;All start times / lanes&rdquo; as ground truth for conflicts, and must confirm Match Location with the Team captain before the match.
    </p>
    <table class="schedule-table">
      <thead>
        <tr>
          <th>Date</th>
          <th>Team name</th>
          <th>Match time</th>
          <th>All start times / lanes</th>
          <th>Opponent team</th>
          <th>Home/Away</th>
          <th>Location<sup><a href="#location-footnote" class="footnote-link">*</a></sup></th>
        </tr>
      </thead>
      <tbody>
        {% for row in schedule %}
        <tr class="{% if row.Is_conflict %}conflict{% endif %}{% if row.Is_pending_schedule %} pending{% endif %}">
          <td>{{ row.Date_display }}</td>
          <td>{{ row['Team name'] }}</td>
          <td>{{ row['Match time'] }}</td>
          <td>
            {% if row.Is_pending_schedule %}
              Schedule not yet posted for this match by Home team. Check again later for updates
            {% else %}
              {{ row['All start times / lanes'] }}
            {% endif %}
          </td>
          <td>{{ row['Opponent team'] }}</td>
          <td class="{% if row.Is_home_match %}homeaway_home{% else %}homeaway_away{% endif %}">{{ row['Home/Away'] }}</td>
          <td class="{% if row.Is_home_match %}location_home{% else %}location_away{% endif %}">{{ row['Location'] }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    <div class="actions">
      <form method="post" action="/download">
        <input type="hidden" name="urls" value="{{ urls_value | e }}">
        <input type="hidden" name="mode" value="{{ mode }}">
        <input type="hidden" name="profile_url" value="{{ profile_url | e }}">
        <input type="hidden" name="schedule_json" value='{{ schedule | tojson | e }}'>
        <button type="submit">Download Excel Schedule</button>
      </form>
    </div>
  </div>
  {% endif %}

  <script>
    function toggleModeInputs() {
      const selected = document.querySelector('input[name="mode"]:checked');
      const mode = selected ? selected.value : 'profile';
      const profileWrap = document.getElementById('profile_url_wrap');
      const urlsWrap = document.getElementById('urls_wrap');

      if (mode === 'profile') {
        profileWrap.style.display = '';
        urlsWrap.style.display = 'none';
      } else {
        profileWrap.style.display = 'none';
        urlsWrap.style.display = '';
      }
    }

    document.querySelectorAll('input[name="mode"]').forEach((el) => {
      el.addEventListener('change', toggleModeInputs);
    });

    // Ensure initial state matches server-rendered `mode`.
    toggleModeInputs();
  </script>
</body>
</html>
"""

time_re = re.compile(r"(\d{1,2}:\d{2}\s*(?:AM|PM))", re.IGNORECASE)

PENDING_SCHEDULE_MESSAGE = 'Schedule not yet posted for this match by Home team. Check again later for updates'

def extract_location_from_all_start_times_cell(cell_text):
    """
    Extract a likely court/facility location from the "All start times / lanes" cell.

    Target examples:
    - "Rinconada Ct 1-3" -> "Rinconada Park"
    - "Cubberley cts 1/2/3" -> "Cubberley Center"
    - "Court 1,2,3 at Bramhall park" -> "Bramhall Park"
    - "Willow Glen Middle School (1399 Curtner Ave)" -> "Willow Glen Middle School"
    - "Buchser middle school courts 3 - 7" -> "Buchser Middle School"
    - "Sunnyvale Tennis Center (La Palmas Courts 1-5)" -> "Sunnyvale Tennis Center"
    """
    if not cell_text:
        return ''

    t = str(cell_text).replace('\u00a0', ' ').strip()
    if not t:
        return ''

    # Remove time strings (may appear multiple times).
    t_wo_times = time_re.sub('', t)
    t_wo_times = re.sub(r'\s+', ' ', t_wo_times).strip()

    def is_valid_location(name):
        """Check if extracted text sounds like a real location."""
        if not name or len(name) < 3:
            return False
        name = name.strip()
        # Must start with uppercase letter (proper noun)
        if not re.match(r'^[A-Z]', name):
            return False
        # Reject common non-location words
        lower = name.lower()
        bad_words = {'and', 'at', 'or', 'the', 'a', 'an', 'in', 'on', 'by', 'for', 'with',
                     'all', 'first', 'second', 'third', 'shift', 'followed', 'available',
                     'warm', 'warmup', 'up', 'courts', 'court', 'ct', 'cts',
                     'no', 'not', 'please', 'thanks', 'thank', 'bring', 'your',
                     'our', 'we', 'will', 'can', 'do', 'if', 'is', 'are', 'it'}
        if lower in bad_words:
            return False
        # Check if it starts with bad words
        if any(lower.startswith(w + ' ') for w in bad_words):
            return False
        # Must have at least one word with 3+ letters
        words = name.split()
        if not any(len(w) >= 3 for w in words):
            return False
        return True

    # Facility-type suffixes with explicit case variations (no (?i) flag
    # so that surrounding patterns can enforce case-sensitivity).
    _fac = r'(?:[Ss]chool|[Cc]enter|[Pp]ark|[Cc]omplex|[Cc]lub|[Ss]tadium|[Aa]rena|[Hh]all|[Ff]ield|[Cc]ourts?)'

    # ── Pattern 1 ──────────────────────────────────────────────────────
    # Find all substrings that end with a facility keyword followed by a
    # delimiter ( , ; . or end-of-string.  Pick the longest valid one.
    # Examples:
    #   "Sunnyvale Tennis Center (La Palmas Courts 1-5)" -> "Sunnyvale Tennis Center"
    #   "Sunnyvale Tennis Center, Courts 1-5" -> "Sunnyvale Tennis Center"
    #   "Carlmont High School Tennis Court, C2-C6" -> "Carlmont High School Tennis Court"
    best = ''
    for m in re.finditer(
        r'([A-Z][A-Za-z]+(?:\s+[A-Za-z]+)*\s+' + _fac + r')\s*[(\,;.]',
        t_wo_times
    ):
        name = re.sub(r'\s+', ' ', m.group(1).strip())
        if is_valid_location(name) and len(name) > len(best):
            best = name
    if best:
        return best.title()

    # ── Pattern 2 ──────────────────────────────────────────────────────
    # "<Name> middle/high school courts <digits>"
    # Example: "Buchser middle school courts 3 - 7" -> "Buchser Middle School"
    # No global (?i): group 1 must start uppercase. School keywords are
    # made case-insensitive with character classes.
    m = re.search(
        r'\b([A-Z][a-zA-Z]+)\s+([Mm]iddle\s+[Ss]chool|[Hh]igh\s+[Ss]chool|[Ee]lementary\s+[Ss]chool)\s+(?:[Tt]ennis\s+)?[Cc]ourts?\b',
        t_wo_times
    )
    if m:
        name = (m.group(1) + ' ' + m.group(2)).strip()
        name = re.sub(r'\s+', ' ', name).title()
        if is_valid_location(name):
            return name

    # ── Pattern 3 ──────────────────────────────────────────────────────
    # "Court(s) <numbers> at <Location>"
    # Example: "Court 1,2,3 at Bramhall park" -> "Bramhall Park"
    # No (?i) — capture group must start with actual uppercase letter.
    m = re.search(
        r'\b[Cc]ourts?\s*[\d/,\\\-\s]+\s+[Aa]t\s+([A-Z][a-z]+(?:\s+[A-Za-z]+)*)',
        t_wo_times
    )
    if m:
        name = m.group(1).strip()
        # Stop at noise words
        name = re.split(r'[\.;]|\b(?:[Ff]irst|[Ss]econd|[Tt]hird|[Ss]hift|[Ff]ollowed|[Aa]vailable|[Ww]armup|[Ww]arm\s+[Uu]p|[Nn]o\s+)\b', name)[0].strip()
        name = re.sub(r'\s+', ' ', name)
        if is_valid_location(name):
            return name.title()

    # ── Pattern 4 ──────────────────────────────────────────────────────
    # "<ProperNoun> Ct <digits>"  (case-sensitive: "Ct" must be capitalized)
    # Example: "Rinconada Ct 1-3" -> "Rinconada Park"
    m = re.search(
        r'\b([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)*)\s+Ct\s+[\d/\\\-]',
        t_wo_times
    )
    if m:
        name = m.group(1).strip()
        if is_valid_location(name):
            if re.search(r'(?i)\b(?:Park|Center|Stadium|Complex|Club|Arena|Field|Courts?)\b', name):
                return name
            return f"{name} Park"

    # ── Pattern 5 ──────────────────────────────────────────────────────
    # "<ProperNoun> cts <digits>"  (case-insensitive "cts")
    # Example: "Cubberley cts 1/2/3" -> "Cubberley Center"
    m = re.search(
        r'\b([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)*)\s+[Cc]ts\s+[\d/\\\-]',
        t_wo_times
    )
    if m:
        name = m.group(1).strip()
        if is_valid_location(name):
            if re.search(r'(?i)\b(?:Park|Center|Stadium|Complex|Club|Arena|Field|Courts?)\b', name):
                return name
            return f"{name} Center"

    # ── Pattern 6 ──────────────────────────────────────────────────────
    # "<ABBREV> Courts <digits>"  (all-caps abbreviation + Courts)
    # Example: "STC Courts 1-5" -> "STC Courts"
    m = re.search(
        r'\b([A-Z]{2,}(?:\s+[A-Z][a-zA-Z]+)*)\s+([Cc]ourts?)\s+[\d/\\\-]',
        t_wo_times
    )
    if m:
        name = m.group(1).strip()
        suffix = m.group(2).strip()
        if is_valid_location(name):
            return f"{name} {suffix.title()}"

    # ── Pattern 7 ──────────────────────────────────────────────────────
    # "at <ABBREV> <digits>"  (all-caps abbreviation directly before digits)
    # Example: "courts are at STC 1-5" -> "STC"
    m = re.search(
        r'\b[Aa]t\s+([A-Z]{2,})\s+[\d/\\\-]',
        t_wo_times
    )
    if m:
        name = m.group(1).strip()
        if is_valid_location(name):
            return name

    return ''


def extract_player_name_from_profile(html):
    """
    Best-effort extraction of the player name from a profile page.
    """
    soup = BeautifulSoup(html, 'html.parser')

    blacklist = {
        'usta', 'northern', 'california', 'norcal', 'login', 'join', 'search', 'calendar',
        'play', 'improve', 'stay current', 'coach', 'organize', 'organize', 'pro tennis',
        'captain', 'email', 'administrator'
    }

    def clean(txt):
        return re.sub(r'\s+', ' ', (txt or '').strip())

    # Common: name appears near the top in <b>/<strong> inside a header table.
    # Example: "Sudipta Biswas"
    candidate_re = re.compile(r"^[A-Z][a-zA-Z.\-']+(?:\s+[A-Z][a-zA-Z.\-']+)+$")

    for tag in soup.find_all(['strong', 'b']):
        txt = clean(tag.get_text(' ', strip=True))
        if not txt:
            continue
        if len(txt) > 60 or len(txt) < 3:
            continue
        low = txt.lower()
        if any(b in low for b in blacklist):
            continue
        if candidate_re.match(txt):
            return txt

    # Fallback: regex across whole page text (first plausible "First Last" name).
    page_text = clean(soup.get_text(' ', strip=True))
    m = re.search(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\b', page_text)
    if m:
        return m.group(1).strip()

    return ''


def extract_home_facility_from_teaminfo(html):
    """
    Extract the "Home Facility" value from a teaminfo page.
    """
    soup = BeautifulSoup(html, 'html.parser')
    text = re.sub(r'\s+', ' ', soup.get_text(' ', strip=True))

    # Example (from teaminfo pages):
    # "... Home facility: Sunnyvale Municipal Tennis Center Web address: NA | Currently playing Local league"
    m = re.search(
        r'(?i)home\s*facility\s*[:\-]?\s*(.+?)\s*(?=web\s*address|currently\s*playing|public\s*notes|captain\b|area\s*[:\-]|\||$)',
        text,
    )
    if m:
        val = m.group(1).strip().strip(' -:;,|')
        return val

    # Fallback: less strict (still try to stop before Web address / Currently playing)
    m = re.search(
        r'(?i)home\s*facility\s*[:\-]?\s*(.+?)\s*(?=web\s*address|currently\s*playing|$)',
        text,
    )
    if m:
        val = m.group(1).strip().strip(' -:;,|')
        return val

    return ''


def extract_team_name(soup, fallback):
    if soup.title and '|' in soup.title.get_text():
        return soup.title.get_text().split('|')[-1].strip()
    h1 = soup.find('h1') or soup.find('h2')
    if h1:
        return h1.get_text(strip=True)
    return fallback


def parse_schedule_html(html, fallback_name, team_page_url, home_facility):
    soup = BeautifulSoup(html, 'html.parser')
    team_name = extract_team_name(soup, fallback_name)

    b_tags = soup.find_all(['b', 'strong'])
    schedule_table = None
    for b in b_tags:
        if 'team schedule' in b.get_text(strip=True).lower():
            for tbl in b.find_all_next('table'):
                header_tr = tbl.find('tr')
                if not header_tr:
                    continue
                header_text = ' '.join(td.get_text(' ', strip=True) for td in header_tr.find_all('td'))
                if 'Match date' in header_text:
                    schedule_table = tbl
                    break
        if schedule_table:
            break

    rows_out = []
    if not schedule_table:
        return rows_out

    for tr in schedule_table.find_all('tr'):
        tds = tr.find_all('td')
        if len(tds) < 9:
            continue
        header_candidate = ' '.join(td.get_text(' ', strip=True) for td in tds[:5])
        if 'Match date' in header_candidate or 'Status' in header_candidate:
            continue
        status = tds[0].get_text(' ', strip=True)
        if not status:
            continue

        date_text = tds[2].get_text(' ', strip=True).replace(' ', '').strip()
        day_text = tds[3].get_text(' ', strip=True)

        time_cell_raw = tds[4].get_text(' ', strip=True)
        full_time = time_cell_raw
        for phrase in ['Last schedule update', 'User:']:
            idx = full_time.find(phrase)
            if idx != -1:
                full_time = full_time[:idx].strip(' .;,')

        m = time_re.search(time_cell_raw)
        if m:
            match_time = m.group(1).upper().replace(' ', '')
        else:
            match_time = ''

        opp_td = tds[5]
        opp_cell = opp_td.get_text(' ', strip=True)
        opponent_team_url = ''
        for a in opp_td.find_all('a', href=True):
            href = a['href']
            if 'teaminfo.asp' in href.lower():
                opponent_team_url = urljoin(team_page_url, href)
                break

        home_away = tds[6].get_text(' ', strip=True)
        is_home_match = home_away.strip().lower().startswith('home')

        # Prefer extracting the actual court/facility from the schedule cell.
        cell_location = extract_location_from_all_start_times_cell(full_time)
        location = cell_location if cell_location else (home_facility if is_home_match else '')
        rows_out.append({
            'Date_raw': date_text,
            'Day': day_text,
            'Team name': team_name,
            'Match time': match_time,
            'All start times / lanes': full_time,
            'Opponent team': opp_cell,
            'Home/Away': home_away,
            'Location': location,
            'Is_home_match': is_home_match,
            'Opponent_team_url': opponent_team_url,
            'Status': status,
        })

        ft_norm = (str(full_time).replace(' ', ' ').strip() if full_time is not None else '')
        ft_lower = ft_norm.lower()
        pending_markers = {'', '-', '—', '–', 'n/a', 'na', 'tbd', 'pending', 'not posted', '.', '...'}
        is_pending = (
            ft_lower in pending_markers
            or ft_lower.startswith('tbd')
            or ft_lower.startswith('pending')
            or 'not posted' in ft_lower
            or ft_lower.startswith('n/a')
            or ft_lower.startswith('na')
        )
        rows_out[-1]['Is_pending_schedule'] = is_pending

    return rows_out


def parse_profile_for_teams(html, base_url):
    """Extract team links from a player profile page.

    Heuristic: look for anchors whose href contains 'teaminfo.asp?id='. Use
    surrounding row text as context to help identify the season/year.
    """
    soup = BeautifulSoup(html, 'html.parser')
    teams = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        if 'teaminfo.asp' in href.lower():
            full_url = urljoin(base_url, href)
            label = a.get_text(' ', strip=True) or 'Team'
            row = a.find_parent('tr')
            context = row.get_text(' ', strip=True) if row else ''
            teams.append({'url': full_url, 'label': label, 'context': context})
    return teams


def build_schedule(urls):
    all_rows = []
    facility_cache = {}

    # Parallelize IO-heavy HTTP fetches to reduce perceived latency.
    # Keep worker count modest to avoid overwhelming the upstream site.
    max_workers = min(8, max(1, len(urls)))

    def fetch_and_parse_team(url):
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        home_facility = extract_home_facility_from_teaminfo(resp.text)
        rows = parse_schedule_html(resp.text, url, team_page_url=url, home_facility=home_facility)
        if not rows:
            raise ValueError(f"Could not find a schedule table on {url}.")
        return url, home_facility, rows

    from concurrent.futures import ThreadPoolExecutor, as_completed

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(fetch_and_parse_team, url): url for url in urls}
        for fut in as_completed(futures):
            url, home_facility, rows = fut.result()
            facility_cache[url] = home_facility
            all_rows.extend(rows)

    if not all_rows:
        raise ValueError("No matches found in the provided pages.")

    # Fill in Location for Away matches using the opponent team's Home facility.
    opponent_urls = {
        r.get('Opponent_team_url')
        for r in all_rows
        if not r.get('Is_home_match') and r.get('Opponent_team_url')
    }
    opponent_urls.discard('')
    missing_opponent_urls = [u for u in opponent_urls if u not in facility_cache]

    if missing_opponent_urls:
        opp_workers = min(8, max(1, len(missing_opponent_urls)))

        def fetch_facility(op_url):
            op_resp = requests.get(op_url, timeout=15)
            op_resp.raise_for_status()
            return op_url, extract_home_facility_from_teaminfo(op_resp.text)

        with ThreadPoolExecutor(max_workers=opp_workers) as executor:
            futures = {executor.submit(fetch_facility, u): u for u in missing_opponent_urls}
            for fut in as_completed(futures):
                op_url, home_facility = fut.result()
                facility_cache[op_url] = home_facility

    for r in all_rows:
        if not r.get('Is_home_match') and not str(r.get('Location', '')).strip():
            op_url = r.get('Opponent_team_url', '')
            r['Location'] = facility_cache.get(op_url, '') if op_url else ''

    df = pd.DataFrame(all_rows)
    df['Date'] = pd.to_datetime(df['Date_raw'], format='%m/%d/%y', errors='coerce')
    df = df.sort_values(by=['Date', 'Match time', 'Team name']).reset_index(drop=True)

    out_df = df[['Date', 'Team name', 'Match time', 'All start times / lanes', 'Opponent team', 'Home/Away', 'Location', 'Is_pending_schedule', 'Is_home_match']].copy()

    # Mark conflicting dates
    counts = out_df['Date'].value_counts(dropna=False)
    conflict_dates = set(counts[counts > 1].index)
    out_df['Is_conflict'] = out_df['Date'].isin(conflict_dates)

    # Date display string, e.g., "Apr 15"
    out_df['Date_display'] = out_df['Date'].dt.strftime('%b %d').str.replace(' 0', ' ', regex=False)

    return out_df


@app.route('/', methods=['GET'])
def index():
    current_year = datetime.now().year
    return render_template_string(
        HTML_TEMPLATE,
        urls_value='',
        profile_url='',
        schedule=None,
        message=None,
        error=False,
        team_choices=None,
        current_year=current_year,
        filtered_to_year=False,
        player_name=None,
        player_first_name=None,
        mode='profile',
    )


@app.route('/generate', methods=['POST'])
def generate():
    urls_raw = request.form.get('urls', '')
    profile_url = request.form.get('profile_url', '').strip()
    selected_team_urls = request.form.getlist('team_urls')
    mode = request.form.get('mode', 'profile')
    player_name = request.form.get('player_name', '').strip()
    player_first_name = None
    current_year = datetime.now().year

    # Stage 2: user selected teams from profile (profile mode only)
    if selected_team_urls:
        urls = selected_team_urls
        try:
            out_df = build_schedule(urls)
        except Exception as e:
            return render_template_string(
                HTML_TEMPLATE,
                message=str(e),
                error=True,
                urls_value='\n'.join(urls),
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=True,
                mode='profile',
                player_name=player_name,
            )

        schedule_df = out_df.copy()
        # Make Date JSON-serializable for the hidden input.
        # The visible table uses Date_display, so it's safe to convert Date here.
        schedule_df['Date'] = schedule_df['Date'].dt.strftime('%m/%d/%y').fillna('')
        if 'Is_pending_schedule' in schedule_df.columns:
            pending_mask = schedule_df['Is_pending_schedule'].fillna(False)
            if 'All start times / lanes' in schedule_df.columns:
                schedule_df.loc[pending_mask, 'All start times / lanes'] = PENDING_SCHEDULE_MESSAGE
        if 'Location' in schedule_df.columns:
            schedule_df['Location'] = schedule_df['Location'].fillna('')
        schedule = schedule_df.to_dict(orient='records')
        return render_template_string(
            HTML_TEMPLATE,
            message=None,
            error=False,
            urls_value='\n'.join(urls),
            profile_url=profile_url,
            schedule=schedule,
            team_choices=None,
            current_year=current_year,
            filtered_to_year=True,
            mode='profile',
            player_name=player_name,
        )

    # Profile mode: fetch profile and present teams to choose
    if mode == 'profile':
        if not profile_url:
            return render_template_string(
                HTML_TEMPLATE,
                message="Please provide a player profile URL.",
                error=True,
                urls_value=urls_raw,
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=False,
                mode='profile',
                player_name=None,
            )
        try:
            resp = requests.get(profile_url, timeout=15)
            resp.raise_for_status()
        except Exception as e:
            return render_template_string(
                HTML_TEMPLATE,
                message=f"Error fetching profile: {e}",
                error=True,
                urls_value=urls_raw,
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=False,
                mode='profile',
                player_name=None,
            )

        teams = parse_profile_for_teams(resp.text, profile_url)
        if not teams:
            return render_template_string(
                HTML_TEMPLATE,
                message="Could not find any team links on that profile.",
                error=True,
                urls_value=urls_raw,
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=False,
                mode='profile',
                player_name=None,
            )

        player_name = extract_player_name_from_profile(resp.text).strip()
        if player_name:
            player_first_name = player_name.split()[0].strip()

        year_str = str(current_year)
        current_teams = [t for t in teams if year_str in t.get('context', '') or year_str in t.get('label', '')]
        if current_teams:
            team_choices = current_teams
            filtered_to_year = True
        else:
            team_choices = teams
            filtered_to_year = False

        return render_template_string(
            HTML_TEMPLATE,
            message=None,
            error=False,
            urls_value=urls_raw,
            profile_url=profile_url,
            schedule=None,
            team_choices=team_choices,
            current_year=current_year,
            filtered_to_year=filtered_to_year,
            mode='profile',
            player_name=player_name if player_name else None,
            player_first_name=player_first_name,
        )

    # Teams mode: use manual team URLs directly
    urls = [u.strip() for u in urls_raw.splitlines() if u.strip()]
    if not urls:
        return render_template_string(
            HTML_TEMPLATE,
            message="Please paste at least one team URL.",
            error=True,
            urls_value=urls_raw,
            profile_url=profile_url,
            schedule=None,
            team_choices=None,
            current_year=current_year,
            filtered_to_year=False,
            mode='teams',
            player_name=None,
        )

    try:
        out_df = build_schedule(urls)
    except Exception as e:
        return render_template_string(
            HTML_TEMPLATE,
            message=str(e),
            error=True,
            urls_value=urls_raw,
            profile_url=profile_url,
            schedule=None,
            team_choices=None,
            current_year=current_year,
            filtered_to_year=False,
            mode='teams',
            player_name=None,
        )

    schedule_df = out_df.copy()
    # Make Date JSON-serializable for the hidden input.
    schedule_df['Date'] = schedule_df['Date'].dt.strftime('%m/%d/%y').fillna('')
    if 'Is_pending_schedule' in schedule_df.columns:
        pending_mask = schedule_df['Is_pending_schedule'].fillna(False)
        if 'All start times / lanes' in schedule_df.columns:
            schedule_df.loc[pending_mask, 'All start times / lanes'] = PENDING_SCHEDULE_MESSAGE
    if 'Location' in schedule_df.columns:
        schedule_df['Location'] = schedule_df['Location'].fillna('')
    schedule = schedule_df.to_dict(orient='records')

    return render_template_string(
        HTML_TEMPLATE,
        message=None,
        error=False,
        urls_value=urls_raw,
        profile_url=profile_url,
        schedule=schedule,
        team_choices=None,
        current_year=current_year,
        filtered_to_year=False,
        mode='teams',
        player_name=None,
    )


@app.route('/download', methods=['POST'])
def download():
    schedule_json = request.form.get('schedule_json', '').strip()

    # Prefer exporting the already-generated schedule (no re-scraping / re-parsing).
    if schedule_json:
        try:
            records = json.loads(schedule_json)
            out_df = pd.DataFrame(records)
            if 'Date' in out_df.columns:
                out_df['Date'] = pd.to_datetime(out_df['Date'], format='%m/%d/%y', errors='coerce')
            if 'Location' not in out_df.columns:
                out_df['Location'] = ''
        except Exception:
            # Fall back to re-generation if JSON export fails for some reason.
            schedule_json = ''

    if not schedule_json:
        urls_raw = request.form.get('urls', '')
        profile_url = request.form.get('profile_url', '').strip()
        mode = request.form.get('mode', 'profile')
        current_year = datetime.now().year

        urls = [u.strip() for u in urls_raw.splitlines() if u.strip()]
        if not urls:
            return render_template_string(
                HTML_TEMPLATE,
                message="Please generate a schedule first.",
                error=True,
                urls_value=urls_raw,
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=False,
                mode=mode,
            )

        try:
            out_df = build_schedule(urls)
        except Exception as e:
            return render_template_string(
                HTML_TEMPLATE,
                message=str(e),
                error=True,
                urls_value=urls_raw,
                profile_url=profile_url,
                schedule=None,
                team_choices=None,
                current_year=current_year,
                filtered_to_year=False,
                mode=mode,
            )

    # Write to in-memory Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        columns_to_export = ['Date', 'Team name', 'Match time', 'All start times / lanes', 'Opponent team', 'Home/Away', 'Location']

        # Ensure the exported text matches the web table:
        # - pending rows show the message instead of the raw "All start times / lanes" value
        if 'Is_pending_schedule' in out_df.columns and 'All start times / lanes' in out_df.columns:
            pending_mask = out_df['Is_pending_schedule'].fillna(False).astype(bool)
            out_df.loc[pending_mask, 'All start times / lanes'] = PENDING_SCHEDULE_MESSAGE

        out_df[columns_to_export].to_excel(writer, index=False, sheet_name='Schedule')
        ws = writer.sheets['Schedule']

        from openpyxl.styles import PatternFill, Font

        # Make header row match the web table header style.
        header_fill = PatternFill(start_color='FF01696F', end_color='FF01696F', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFFFF')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        # Column widths + wrapping to avoid manual resizing.
        # Widths are fixed to keep the Excel view consistent.
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        col_widths = {
            'Date': 10,
            'Team name': 28,
            'Match time': 14,
            'All start times / lanes': 70,
            'Opponent team': 28,
            'Home/Away': 10,
            'Location': 45,
        }
        for i, name in enumerate(columns_to_export, start=1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(name, 20)

        # Use the same alignment for all cells to avoid mixed look.
        uniform_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for excel_row in range(1, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=c_idx).alignment = uniform_alignment

        # Format date column as "mmm d"
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if isinstance(cell.value, datetime):
                cell.number_format = 'mmm d'

        # Styling to match the web table
        pending_fill = PatternFill(start_color='FFEFEFEF', end_color='FFEFEFEF', fill_type='solid')
        conflict_fill = PatternFill(start_color='FFFFF7C2', end_color='FFFFF7C2', fill_type='solid')

        pending_font = Font(color='FF9A9A9A', italic=True)
        conflict_base_font = Font(color='FF222222')

        homeaway_home_font = Font(color='FF01696F', bold=True)
        homeaway_away_font = Font(color='FFA12C2C', bold=True)
        location_home_font = Font(color='FF0B6B0B', bold=True)
        location_away_font = Font(color='FF6B2D91', bold=True)

        pending_mask = out_df['Is_pending_schedule'].fillna(False).astype(bool) if 'Is_pending_schedule' in out_df.columns else None
        is_home_mask = out_df['Is_home_match'].fillna(True).astype(bool) if 'Is_home_match' in out_df.columns else None

        conflict_mask = out_df['Date'].duplicated(keep=False) if 'Date' in out_df.columns else None

        homeaway_col_idx = columns_to_export.index('Home/Away') + 1
        location_col_idx = columns_to_export.index('Location') + 1
        col_count = len(columns_to_export)
        # Match the web table header: "Location *" (superscript in HTML)
        ws.cell(row=1, column=location_col_idx).value = 'Location *'

        for i in range(len(out_df)):
            excel_row = i + 2
            is_pending = bool(pending_mask.iloc[i]) if pending_mask is not None else False
            is_conflict = bool(conflict_mask.iloc[i]) if conflict_mask is not None else False
            is_home = bool(is_home_mask.iloc[i]) if is_home_mask is not None else True

            if is_pending:
                # Pending overrides everything else (including yellow conflict fill)
                for c in range(1, col_count + 1):
                    cell = ws.cell(row=excel_row, column=c)
                    cell.fill = pending_fill
                    cell.font = pending_font
                continue

            if is_conflict:
                for c in range(1, col_count + 1):
                    cell = ws.cell(row=excel_row, column=c)
                    cell.fill = conflict_fill
                    cell.font = conflict_base_font

            # Apply home/away + location colors (matches CSS with !important rules)
            ws.cell(row=excel_row, column=homeaway_col_idx).font = homeaway_home_font if is_home else homeaway_away_font
            ws.cell(row=excel_row, column=location_col_idx).font = location_home_font if is_home else location_away_font

    output.seek(0)
    filename = 'usta_schedule_combined.xlsx'
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True)
